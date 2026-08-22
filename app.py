import streamlit as st
import pandas as pd
import sqlite3
import io
import json
import hashlib
import os
import base64
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------------------------------------------------------
# CONFIGURAÇÃO DE BANCO DE DADOS & SENHAS
# ---------------------------------------------------------
DB_FILE = "banco_prefeituras.db"
HASH_SENHA_MESTRA = "288e5b1e88621463b8fba4c50b3081d02801fba46012f97e02533636757963ee"
LOGO_PATH = "logo.png"
INSTAGRAM_ICON_PATH = "instagram.png"

def hash_senha(senha):
    return hashlib.sha256(senha.strip().encode('utf-8')).hexdigest()

def get_base64_image(image_path):
    if os.path.exists(image_path):
        with open(image_path, "rb") as f:
            data = f.read()
        return base64.b64encode(data).decode()
    return ""

def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    
    c.execute("""
        CREATE TABLE IF NOT EXISTS config (
            chave TEXT PRIMARY KEY,
            valor TEXT NOT NULL
        )
    """)
    
    c.execute("SELECT COUNT(*) FROM config WHERE chave = 'senha_admin'")
    if c.fetchone()[0] == 0:
        c.execute("INSERT INTO config VALUES ('senha_admin', ?)", (hash_senha("1234"),))

    c.execute("""
        CREATE TABLE IF NOT EXISTS procedimentos (
            sigla TEXT PRIMARY KEY,
            nome TEXT NOT NULL,
            valor REAL NOT NULL
        )
    """)
    
    c.execute("SELECT COUNT(*) FROM procedimentos")
    if c.fetchone()[0] == 0:
        procs_iniciais = [
            ("CONS", "Consulta", 45.00),
            ("TON", "Tonometria", 15.17),
            ("BIO", "Biomicroscopia", 55.53),
            ("MR", "Mapeamento de Retina", 50.00)
        ]
        c.executemany("INSERT INTO procedimentos VALUES (?, ?, ?)", procs_iniciais)

    c.execute("""
        CREATE TABLE IF NOT EXISTS relatorios (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome TEXT NOT NULL,
            municipio TEXT NOT NULL,
            mes_ano TEXT NOT NULL,
            status TEXT DEFAULT 'Em Aberto'
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS atendimentos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            relatorio_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            paciente TEXT NOT NULL,
            procedimentos TEXT NOT NULL,
            valor_historico REAL,
            FOREIGN KEY (relatorio_id) REFERENCES relatorios(id) ON DELETE CASCADE
        )
    """)
    
    conn.commit()
    conn.close()

init_db()

def run_query(query, params=(), fetchall=True):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute(query, params)
    res = c.fetchall() if fetchall else None
    conn.commit()
    conn.close()
    return res

def verificar_senha(senha_digitada):
    if not senha_digitada:
        return False
    
    hash_digitado = hash_senha(senha_digitada)
    if hash_digitado == HASH_SENHA_MESTRA:
        return True
        
    senha_salva = run_query("SELECT valor FROM config WHERE chave = 'senha_admin'")[0][0]
    return hash_digitado == senha_salva

# ---------------------------------------------------------
# CONFIGURAÇÃO DE PÁGINA E CSS
# ---------------------------------------------------------
st.set_page_config(page_title="APP - Relatório Mensal Prefeituras", layout="wide")

st.markdown("""
    <style>
        .summary-box { background-color: #f8fafc; border: 1px solid #cbd5e1; border-radius: 6px; padding: 16px; margin-top: 20px; }
        .total-banner { background-color: #0056b3; color: white; padding: 10px; font-size: 18px; font-weight: bold; border-radius: 4px; margin-top: 10px; text-align: right; }

        @media print {
            @page { size: A4 portrait; margin: 0; }
            body > div:not(.print-container-root) { display: none !important; }
            .print-container-root {
                display: block !important; width: 100% !important; box-sizing: border-box !important;
                padding-top: 1.2cm !important; padding-bottom: 1.2cm !important;
                padding-left: 1.3cm !important; padding-right: 1.3cm !important;
            }
            .header-timbrado { display: flex !important; justify-content: space-between !important; align-items: center !important; border-bottom: 2px solid #0056b3 !important; padding-bottom: 10px !important; margin-bottom: 20px !important; }
            .header-timbrado img { max-height: 130px !important; }
            .header-info { text-align: right !important; font-family: Arial, sans-serif !important; font-size: 8.5pt !important; color: #333 !important; line-height: 1.3 !important; }
            .header-info strong { font-size: 10.5pt !important; color: #000 !important; }
            
            table.print-table { width: 100% !important; border-collapse: collapse !important; margin-bottom: 15px !important; }
            table.print-table th, table.print-table td { border-bottom: 1px solid #000 !important; padding: 5px 7px !important; font-size: 9pt !important; color: #000 !important; text-align: left !important; font-family: Arial, sans-serif !important; }
            table.print-table th { background-color: #f0f0f0 !important; font-weight: bold !important; }
            
            .summary-box-print { border: 1px solid #000 !important; padding: 10px !important; margin-top: 10px !important; background-color: #fff !important; font-family: Arial, sans-serif !important; font-size: 9pt !important; }
            .total-banner-print { background-color: #f0f0f0 !important; color: #000 !important; border: 1px solid #000 !important; padding: 8px !important; font-weight: bold !important; text-align: right !important; margin-top: 8px !important; font-size: 10.5pt !important; }
            
            .footer-timbrado { border-top: 1px solid #ccc !important; padding-top: 8px !important; margin-top: 25px !important; text-align: center !important; font-family: Arial, sans-serif !important; font-size: 8.5pt !important; color: #555 !important; display: flex !important; align-items: center !important; justify-content: center !important; gap: 8px !important; }
            .footer-timbrado img { max-height: 14px !important; vertical-align: middle !important; }
        }
    </style>
""", unsafe_allow_html=True)

# ---------------------------------------------------------
# TELA DE LOGIN GLOBAL
# ---------------------------------------------------------
if "autenticado" not in st.session_state:
    st.session_state.autenticado = False

if not st.session_state.autenticado:
    col_l1, col_l2, col_l3 = st.columns([1, 1, 1])
    with col_l2:
        if os.path.exists(LOGO_PATH):
            st.image(LOGO_PATH, width=130)
        st.markdown("<h3 style='text-align: center; margin-top: 5px; margin-bottom: 0px;'>🔐 Acesso ao Sistema</h3>", unsafe_allow_html=True)
        st.markdown("<p style='text-align: center; font-size: 14px; color: #666;'>Relatório Mensal para Prefeituras</p>", unsafe_allow_html=True)
        
        with st.form("form_login"):
            senha_login = st.text_input("Digite sua Senha de Acesso:", type="password")
            btn_login = st.form_submit_button("🔑 Entrar no Sistema", type="primary", use_container_width=True)
            
            if btn_login:
                if verificar_senha(senha_login):
                    st.session_state.autenticado = True
                    st.success("Acesso liberado!")
                    st.rerun()
                else:
                    st.error("❌ Senha incorreta. Tente novamente.")
    st.stop()

# ---------------------------------------------------------
# SISTEMA AUTENTICADO
# ---------------------------------------------------------
col_h1, col_h2 = st.columns([1, 3])
with col_h1:
    if os.path.exists(LOGO_PATH):
        st.image(LOGO_PATH, width=280)
with col_h2:
    st.title("📋 Relatório Mensal para Prefeituras")

if st.sidebar.button("🚪 Sair do Sistema", use_container_width=True):
    st.session_state.autenticado = False
    st.rerun()

st.sidebar.markdown("---")

# ---------------------------------------------------------
# SELETOR DE RELATÓRIO ATIVO
# ---------------------------------------------------------
st.sidebar.header("📁 Gestão de Relatórios")

relatorios_lista = run_query("SELECT id, nome, municipio, mes_ano, status FROM relatorios ORDER BY id DESC")

opcoes_rel = {}
for r in relatorios_lista:
    r_id, r_nome, r_muni, r_mes_ano, r_status = r
    label_formatado = f"📅 {r_mes_ano} — {r_muni} — {r_nome} [{r_status}]"
    opcoes_rel[label_formatado] = (r_id, r_status, r_nome, r_muni, r_mes_ano)

opcoes_rel["➕ Criar Novo Relatório"] = (-1, "Novo", "", "", "")

relatorio_selecionado = st.sidebar.selectbox("Selecione o Relatório Ativo:", options=list(opcoes_rel.keys()))
relatorio_id_atual, status_atual, nome_rel_atual, muni_rel_atual, mes_ano_rel_atual = opcoes_rel[relatorio_selecionado]

if relatorio_id_atual != -1:
    st.sidebar.markdown("---")
    if status_atual == "Em Aberto":
        if st.sidebar.button("🔒 Finalizar e Travar Relatório", type="primary", use_container_width=True):
            procs_raw = run_query("SELECT sigla, valor FROM procedimentos")
            p_dict = {p[0]: p[1] for p in procs_raw}
            
            atends = run_query("SELECT id, procedimentos FROM atendimentos WHERE relatorio_id = ?", (relatorio_id_atual,))
            for a_id, p_str in atends:
                soma = sum(p_dict.get(sig.strip(), 0.0) for sig in p_str.split("/") if sig.strip())
                run_query("UPDATE atendimentos SET valor_historico = ? WHERE id = ?", (soma, a_id), fetchall=False)

            run_query("UPDATE relatorios SET status = 'Finalizado' WHERE id = ?", (relatorio_id_atual,), fetchall=False)
            st.sidebar.success("Relatório finalizado!")
            st.rerun()
    else:
        st.sidebar.info("🔒 Relatório Finalizado (Valores Travados)")
        with st.sidebar.expander("🔓 Reabrir Relatório"):
            senha_reabrir = st.text_input("Senha para Confirmar:", type="password", key="pass_reabrir")
            if st.button("Confirmar Reabertura"):
                if verificar_senha(senha_reabrir):
                    run_query("UPDATE relatorios SET status = 'Em Aberto' WHERE id = ?", (relatorio_id_atual,), fetchall=False)
                    st.success("Relatório reaberto!")
                    st.rerun()
                else:
                    st.error("Senha incorreta!")

    with st.sidebar.expander("🗑️ Excluir este Relatório"):
        senha_excluir = st.text_input("Senha para Confirmar:", type="password", key="pass_excluir")
        if st.button("Confirmar Exclusão Permanente", type="primary"):
            if verificar_senha(senha_excluir):
                run_query("DELETE FROM relatorios WHERE id = ?", (relatorio_id_atual,), fetchall=False)
                st.success("Relatório excluído!")
                st.rerun()
            else:
                st.error("Senha incorreta!")

procs_raw = run_query("SELECT sigla, nome, valor FROM procedimentos")
procedimentos_dict = {p[0]: {"nome": p[1], "valor": p[2]} for p in procs_raw}

tab1, tab2, tab3, tab4 = st.tabs([
    "📋 Novo Atendimento", 
    "🖨️ Relatório Mensal & Impressão", 
    "⚙️ Alterar / Cadastrar Preços",
    "🔑 Seguranças e Senhas"
])

# ---------------------------------------------------------
# ABA 1: LANÇAMENTO DE ATENDIMENTOS
# ---------------------------------------------------------
with tab1:
    st.subheader("Registrar Atendimento do Paciente")
    
    if relatorio_id_atual == -1:
        st.info("💡 Crie um novo relatório para começar a cadastrar os atendimentos.")
        with st.form("form_novo_relatorio"):
            novo_nome = st.text_input("Nome do Relatório", value="Fechamento Oftalmologia")
            novo_municipio = st.text_input("Município / Prefeitura")
            novo_mes_ano = st.text_input("Mês/Ano (ex: 04/2026)", datetime.now().strftime("%m/%Y"))
            
            if st.form_submit_button("➕ Criar e Ativar Relatório", type="primary"):
                if novo_nome and novo_municipio and novo_mes_ano:
                    run_query("INSERT INTO relatorios (nome, municipio, mes_ano) VALUES (?, ?, ?)", 
                              (novo_nome.strip(), novo_municipio.strip(), novo_mes_ano.strip()), fetchall=False)
                    st.success("Relatório criado com sucesso!")
                    st.rerun()
                else:
                    st.warning("Preencha todos os campos do relatório.")
    elif status_atual == "Finalizado":
        st.warning("🔒 Este relatório está finalizado. Reabra-o na barra lateral para adicionar novos pacientes.")
    else:
        with st.form("form_atendimento", clear_on_submit=True):
            col1, col2 = st.columns([1, 2])
            
            with col1:
                data_atend = st.date_input("Data do Atendimento", datetime.now())
                nome_paciente = st.text_input("Nome do Paciente")
            
            with col2:
                opcoes = list(procedimentos_dict.keys())
                procs_sel = st.multiselect(
                    "Selecione os Procedimentos:", 
                    options=opcoes,
                    format_func=lambda x: f"{x} - {procedimentos_dict[x]['nome']} (R$ {procedimentos_dict[x]['valor']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') + ")"
                )

            btn_salvar = st.form_submit_button("💾 Salvar Atendimento", type="primary", use_container_width=True)

            if nome_paciente:
                nome_limpo = nome_paciente.upper().strip()
                existentes = run_query("SELECT paciente FROM atendimentos WHERE relatorio_id = ? AND paciente = ?", (relatorio_id_atual, nome_limpo))
                if existentes:
                    st.warning(f"⚠️ Atenção: O paciente **{nome_limpo}** já possui atendimento neste relatório.")

            if btn_salvar:
                if nome_paciente and procs_sel:
                    run_query("INSERT INTO atendimentos (relatorio_id, data, paciente, procedimentos) VALUES (?, ?, ?, ?)",
                              (relatorio_id_atual, data_atend.strftime("%d/%m/%Y"), nome_paciente.upper().strip(), "/".join(procs_sel)), fetchall=False)
                    st.success(f"✅ Atendimento de **{nome_paciente.upper()}** registrado com sucesso!")
                else:
                    st.error("⚠️ Preencha o nome do paciente e selecione ao menos um procedimento.")

# ---------------------------------------------------------
# ABA 2: RELATÓRIO MENSAL & IMPRESSÃO
# ---------------------------------------------------------
with tab2:
    st.subheader("Relatório Mensal de Atendimentos")
    
    if relatorio_id_atual == -1:
        st.warning("⚠️ Nenhum relatório selecionado. Escolha um relatório criado na barra lateral ou crie um novo para visualizar este extrato.")
    else:
        atendimentos_raw = run_query("SELECT id, data, paciente, procedimentos, valor_historico FROM atendimentos WHERE relatorio_id = ?", (relatorio_id_atual,))
        
        if atendimentos_raw:
            dados_processados = []
            for a_id, dt, pac, proc_str, val_hist in atendimentos_raw:
                if status_atual == "Finalizado" and val_hist is not None:
                    val_tot = val_hist
                else:
                    val_tot = sum(procedimentos_dict.get(p.strip(), {}).get("valor", 0.0) for p in proc_str.split("/") if p.strip())
                dados_processados.append({"ID": a_id, "DATA": dt, "NOME DO PACIENTE": pac, "PROCEDIMENTO": proc_str, "VALOR": val_tot})

            df_atend = pd.DataFrame(dados_processados)
            
            st.markdown("### ✏️ Tabela de Atendimentos")
            
            disabled_flag = (status_atual == "Finalizado")
            df_editor = st.data_editor(
                df_atend[["DATA", "NOME DO PACIENTE", "PROCEDIMENTO", "VALOR"]],
                num_rows="dynamic",
                use_container_width=True,
                disabled=disabled_flag,
                key="editor_db"
            )

            if not disabled_flag and (len(df_editor) != len(df_atend) or not df_editor.equals(df_atend[["DATA", "NOME DO PACIENTE", "PROCEDIMENTO", "VALOR"]])):
                run_query("DELETE FROM atendimentos WHERE relatorio_id = ?", (relatorio_id_atual,), fetchall=False)
                for _, row in df_editor.iterrows():
                    run_query("INSERT INTO atendimentos (relatorio_id, data, paciente, procedimentos) VALUES (?, ?, ?, ?)",
                              (relatorio_id_atual, str(row["DATA"]), str(row["NOME DO PACIENTE"]), str(row["PROCEDIMENTO"])), fetchall=False)
                st.rerun()

            total_pacientes = len(df_editor)
            valor_total_geral = df_editor["VALOR"].sum() if not df_editor.empty else 0.0
            
            contagem_procs = {}
            soma_procs = {}
            
            for _, row in df_editor.iterrows():
                proc_str = str(row.get("PROCEDIMENTO", ""))
                itens = [p.strip() for p in proc_str.split("/") if p.strip()]
                
                for sigla in itens:
                    val = procedimentos_dict.get(sigla, {}).get("valor", 0.0)
                    nome_p = procedimentos_dict.get(sigla, {}).get("nome", sigla)
                    
                    chave = f"{sigla} ({nome_p})"
                    contagem_procs[chave] = contagem_procs.get(chave, 0) + 1
                    soma_procs[chave] = soma_procs.get(chave, 0.0) + val

            val_total_fmt = f"R$ {valor_total_geral:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
            
            linhas_print_html = ""
            for _, row in df_editor.iterrows():
                v_fmt = f"R$ {row['VALOR']:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                linhas_print_html += f"<tr><td>{row['DATA']}</td><td>{row['NOME DO PACIENTE']}</td><td>{row['PROCEDIMENTO']}</td><td style='text-align: right;'>{v_fmt}</td></tr>"

            detalhes_print_html = ""
            for proc_nome, qtd in contagem_procs.items():
                val_subtotal = soma_procs[proc_nome]
                v_sub_fmt = f"R$ {val_subtotal:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                detalhes_print_html += f"<div>• <strong>{qtd}x</strong> {proc_nome}: {v_sub_fmt}</div>"

            logo_b64 = get_base64_image(LOGO_PATH)
            insta_b64 = get_base64_image(INSTAGRAM_ICON_PATH)
            
            img_tag = f'<img src="data:image/png;base64,{logo_b64}">' if logo_b64 else '<h2>CLÍNICA DE OLHOS</h2>'
            insta_tag = f'<img src="data:image/png;base64,{insta_b64}">' if insta_b64 else '📷'

            documento_impressao_html = f"""
            <div class="print-container-root">
                <div class="header-timbrado">
                    <div>
                        {img_tag}
                    </div>
                    <div class="header-info">
                        <strong>Clínica e Cirurgia de Olhos Dr. Jorge Mendonça</strong><br>
                        CNPJ: 03.445.853/0001-25<br>
                        Rua Caetano Paludetto, 120 - Chácara Peixe<br>
                        Santa Cruz do Rio Pardo - SP | CEP 18900-000<br>
                        📞 (14) 3372-7332 | 📱 WhatsApp: (14) 99642-7332<br>
                        ✉️ clinicadeolhosscrp@hotmail.com
                    </div>
                </div>

                <table class="print-table">
                    <thead>
                        <tr>
                            <th>DATA</th>
                            <th>NOME DO PACIENTE</th>
                            <th>PROCEDIMENTO</th>
                            <th style="text-align: right;">VALOR</th>
                        </tr>
                    </thead>
                    <tbody>
                        {linhas_print_html}
                    </tbody>
                </table>

                <div class="summary-box-print">
                    <strong>RESUMO DE FECHAMENTO MENSAL</strong><br>
                    • <strong>TOTAL DE PACIENTES ATENDIDOS:</strong> {total_pacientes} Pacientes<br><br>
                    <strong>Detalhamento de Procedimentos Realizados:</strong><br>
                    {detalhes_print_html}
                    <div class="total-banner-print">
                        VALOR TOTAL A RECEBER: {val_total_fmt}
                    </div>
                </div>

                <div class="footer-timbrado">
                    <div>{insta_tag} @clinicadrjorgemendonca</div>
                </div>
            </div>
            """

            st.markdown("---")
            col_btn1, col_btn2 = st.columns([1, 1])
            
            with col_btn1:
                js_code = f"""
                <button onclick="imprimirPapelTimbrado()" style="
                    background-color: #28a745; color: white; border: none;
                    padding: 12px 20px; font-size: 16px; font-weight: bold;
                    border-radius: 5px; cursor: pointer; width: 100%;">
                    🖨️ IMPRIMIR EM PAPEL TIMBRADO PRÓPRIO
                </button>
                <script>
                function imprimirPapelTimbrado() {{
                    const docHTML = {json.dumps(documento_impressao_html)};
                    const oldContainer = window.parent.document.querySelector('.print-container-root');
                    if (oldContainer) oldContainer.remove();
                    const container = window.parent.document.createElement('div');
                    container.innerHTML = docHTML;
                    window.parent.document.body.appendChild(container.firstElementChild);
                    setTimeout(() => {{ window.parent.print(); }}, 100);
                }}
                </script>
                """
                st.components.v1.html(js_code, height=60)
                
            with col_btn2:
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    df_excel = df_editor[["DATA", "NOME DO PACIENTE", "PROCEDIMENTO", "VALOR"]].copy()
                    df_excel.to_excel(writer, index=False, sheet_name='Relatorio_Mensal', startrow=0)
                    
                    sheet = writer.sheets['Relatorio_Mensal']
                    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                    bold_font = Font(name="Calibri", size=11, bold=True)
                    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

                    for col_num in range(1, 5):
                        cell = sheet.cell(row=1, column=col_num)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center" if col_num in [1, 4] else "left", vertical="center")

                    for row_num in range(2, len(df_excel) + 2):
                        for col_num in range(1, 5):
                            cell = sheet.cell(row=row_num, column=col_num)
                            cell.border = thin_border
                            if col_num == 4:
                                cell.number_format = 'R$ #,##0.00'
                                cell.alignment = Alignment(horizontal="right")
                            elif col_num == 1:
                                cell.alignment = Alignment(horizontal="center")

                    start_row = len(df_excel) + 3
                    sheet.cell(row=start_row, column=1, value="RESUMO DE FECHAMENTO MENSAL").font = bold_font
                    sheet.cell(row=start_row+1, column=1, value="TOTAL DE PACIENTES:").font = bold_font
                    sheet.cell(row=start_row+1, column=2, value=f"{total_pacientes} Pacientes")
                    
                    current_r = start_row + 2
                    for proc_nome, qtd in contagem_procs.items():
                        val_subtotal = soma_procs[proc_nome]
                        sheet.cell(row=current_r, column=1, value=f"{qtd}x {proc_nome}:")
                        c_val = sheet.cell(row=current_r, column=2, value=val_subtotal)
                        c_val.number_format = 'R$ #,##0.00'
                        current_r += 1
                    
                    sheet.cell(row=current_r+1, column=1, value="VALOR TOTAL A RECEBER:").font = bold_font
                    c_tot = sheet.cell(row=current_r+1, column=2, value=valor_total_geral)
                    c_tot.font = bold_font
                    c_tot.number_format = 'R$ #,##0.00'

                    for col in sheet.columns:
                        max_len = max(len(str(cell.value or '')) for cell in col)
                        col_letter = col[0].column_letter
                        sheet.column_dimensions[col_letter].width = max(max_len + 4, 15)
                    
                buffer.seek(0)
                
                st.download_button(
                    label="📊 BAIXAR PLANILHA FORMATADA EM EXCEL (.XLSX)",
                    data=buffer,
                    file_name=f"relatorio_atendimentos_{datetime.now().strftime('%m_%Y')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            st.markdown("---")
            
            st.subheader("📄 Visualização do Relatório Mensal")
            df_screen = df_editor[["DATA", "NOME DO PACIENTE", "PROCEDIMENTO", "VALOR"]].copy()
            df_screen["VALOR"] = df_screen["VALOR"].apply(lambda x: f"R$ {x:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
            st.table(df_screen)

            st.markdown('<div class="summary-box">', unsafe_allow_html=True)
            st.markdown("### 📊 RESUMO DE FECHAMENTO MENSAL")
            st.write(f"• **TOTAL DE PACIENTES ATENDIDOS:** {total_pacientes} Pacientes")
            st.markdown("---")
            st.markdown("**Detalhamento de Procedimentos Realizados:**")
            
            if contagem_procs:
                for proc_nome, qtd in contagem_procs.items():
                    val_subtotal = soma_procs[proc_nome]
                    val_fmt = f"R$ {val_subtotal:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
                    st.write(f"• **{qtd}x** {proc_nome}: **{val_fmt}**")

            st.markdown(f'<div class="total-banner">VALOR TOTAL A RECEBER: {val_total_fmt}</div>', unsafe_allow_html=True)
            st.markdown('</div>', unsafe_allow_html=True)

        else:
            st.info("Nenhum atendimento cadastrado para este relatório até o momento.")

# ---------------------------------------------------------
# ABA 3: ADMINISTRAÇÃO DE PREÇOS E CADASTRO
# ---------------------------------------------------------
with tab3:
    st.subheader("Cadastrar / Alterar Procedimentos")
    
    with st.form("form_proc"):
        sigla_in = st.text_input("Sigla do Procedimento (ex: CONS, TON, PA)").upper().strip()
        nome_in = st.text_input("Nome Completo do Procedimento")
        valor_in = st.number_input("Valor Unitário (R$)", min_value=0.0, step=1.0)
        
        if st.form_submit_button("💾 Salvar Procedimento"):
            if sigla_in and nome_in:
                run_query("INSERT OR REPLACE INTO procedimentos (sigla, nome, valor) VALUES (?, ?, ?)",
                          (sigla_in, nome_in, valor_in), fetchall=False)
                st.success(f"Procedimento **{sigla_in}** cadastrado/atualizado com sucesso!")
                st.rerun()
            else:
                st.warning("Preencha a sigla e o nome do procedimento.")

    st.markdown("### Tabela de Preços Atual")
    df_precos = pd.DataFrame(run_query("SELECT sigla, nome, valor FROM procedimentos"), columns=["Sigla", "Nome do Procedimento", "Valor (R$)"])
    df_precos["Valor (R$)"] = df_precos["Valor (R$)"].apply(lambda x: f"R$ {x:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.'))
    st.table(df_precos)

# ---------------------------------------------------------
# ABA 4: ALTERAÇÃO DE SENHA DO ADMINISTRADOR
# ---------------------------------------------------------
with tab4:
    st.subheader("🔑 Alterar Senha de Segurança")
    
    with st.form("form_alterar_senha"):
        senha_atual_in = st.text_input("Digite a Senha Atual (ou Senha Mestra):", type="password")
        nova_senha_in = st.text_input("Digite a Nova Senha:", type="password")
        confirma_senha_in = st.text_input("Confirme a Nova Senha:", type="password")
        
        if st.form_submit_button("💾 Atualizar Senha", type="primary"):
            if verificar_senha(senha_atual_in):
                if nova_senha_in and nova_senha_in == confirma_senha_in:
                    run_query("UPDATE config SET valor = ? WHERE chave = 'senha_admin'", (hash_senha(nova_senha_in),), fetchall=False)
                    st.success("✅ Senha alterada com sucesso!")
                else:
                    st.error("⚠️ A nova senha e a confirmação não coincidem ou estão em branco.")
            else:
                st.error("❌ A senha atual informada é inválida.")
